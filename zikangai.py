import openpyxl as px
import streamlit as st

st.title('時間外記入アプリ')
filepath = st.file_uploader("1．勤務表をアップロードしてください。", type="xlsx")
button = None
if filepath:
    wb = px.load_workbook(filepath)
    sheet_names = wb.sheetnames
    st.write(f'シート一覧:{sheet_names}')

    left_column, right_column = st.columns(2)

    year = left_column.selectbox('年を選択', ('2022', '2023', '2024', '2025', '2026'))
    year = int(year) % 100
    month = right_column.selectbox('月を選択', ('01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12'))
    wb = px.load_workbook(filepath)
    sheet_name1 = str(year)+'.'+str(month)
    sheet = wb[sheet_name1]

    member = []
    for i in range(59):
        if sheet.cell(row=4+i, column=2).value == '夜勤':
            break
        member.append(sheet.cell(row=4+i, column=2).value.replace(' ', '').replace('　', ''))
    # st.write(f'{member}')


    name = st.selectbox('名前を選択', ('安藤　雅','浅野　和也','小鹿野奈緒美','板橋　佑典','小出　純一','横山　絢香','宮島　優実','浅井　勇太','成田　瑞生','福田　淳也','坂井　義行','宮澤　仁美','中村　翔','佐藤　有将','日向　真悟','竹内　友一','関　 優子','長浜　大輔','都丸　浩美','松井　聡美','氏田　浩一','品川　博史','尾崎　大輔','黒澤　裕司','須田　浩太','嶋田　博孝','岸　和洋','岡田　良介','新井　啓祐','茂木　直','樋口　弘光','小屋　順一','松村　直樹','幅野　陽二','岡田　大希','及川　聡子','丹　章吾','安部　聖','中村　潤平','鑓田　和真','星野　佳彦','小鹿野友昭','津田　和寿','勘崎　貴雄','宮澤　真','大橋　慶明','米原　絵理','湯浅　大智','小野　将平','山口　直人','村中　愛美','田子　智也','青木　颯斗','増田　洋明'))
    name = str(name).replace(' ', '').replace('　', '')

    myRanges = sheet['B3:B62'] #勤務表専用の範囲
    # st.write(name)

    num = []
    for r in myRanges:
        for cell in r:
          num.append(cell.coordinate)
        #   print(str(cell.value).replace(' ', '').replace('　', ''))
          if name in str(cell.value).replace(' ', '').replace('　', ''):
              num_moji = str(cell.value).replace(' ', '').replace('　', '') #名前(日向　真悟、日向、真悟)
              num_key = cell.coordinate #セル番号(ex:B18)
            #   print(num_moji,num_key)

    #日付をもとに終わりの列を取得(max_clm)
    for i in range(10,40):
      if sheet.cell(row=1, column=i).value == None:
        max_clm = i - 1
        # print('終点column:%d' % max_clm)
        break
    #日付をもとに始まりの列を取得(min_clm)
    for i in range(10, 0, -1):
      if sheet.cell(row=1, column=i).value == 1:
        min_clm = i
        # print('始点column:%d' % min_clm)

    #目的Cellの行列を取得
    r_t = sheet[num_key].row     #名前を含むCell:B18 のrow
    c_t = sheet[num_key].column  #名前を含むCell:B18　のcolumn
    # print(r_t, c_t)

    #目的の行を取得
    main = []
    for i in range(min_clm,max_clm+1):
      main.append(sheet.cell(row=r_t, column=i).value)

    filepath_2 = st.file_uploader("2．記載したい時間外ファイルをアップロードしてください。", type="xlsx")
    if filepath_2:
        wb = px.load_workbook(filepath_2)
        sheet_name2 = str(int(month))+'月'
        sheet_2 = wb[sheet_name2]
        st.write(f'シート名：{sheet_name2}')

    button = st.button('書込開始')
#メインの計算
if button:
    #勤務表への書込み

    start = 6 #時間外シートの開始位置
    n = max_clm-min_clm+1
    print('日数確認用：%d' % n) #日数

    #書込み
    for i in range(n):
        a = b = c = d = e = f = None
        x = main[i]
        #print(x)
        if x == None or x == 'AG' or x == '重' or x == '治' or x == '/講' or x == 'RI':
            a, b, c, d, e, f = 8, 30, 17, 15, 60, None
        elif x == 'A' or x == 'ICU':
            a, b, c, d, e, f = 8, 30, 17, 15, 60, '日勤'
        elif x == '★':
            a, b, c, d, e, f = 16, 30, 9, 30, 120, '夜勤'
        elif x == '☆':
            a, b, c, d, e, f = None, None, None, None, None, '非番'
        elif x == '/':
            a, b, c, d, e, f = None, None, None, None, None, None
        elif x == '×':
            a, b, c, d, e, f = None, None, None, None, None, '振替休日(/)'
        elif x == '半有':
            a, b, c, d, e, f = 8, 30, 17, 15, 60, '午後休暇'
        elif x == '年':
            a, b, c, d, e, f = 8, 30, 17, 15, 60, '有給休暇'
        elif x == '年1h'or x == '特1h':
            a, b, c, d, e, f = 8, 30, 17, 15, 60, '1時間休'
        elif x == '年2h':
            a, b, c, d, e, f = 8, 30, 17, 15, 60, '2時間休'
        elif x == '特':
            a, b, c, d, e, f = 8, 30, 17, 15, 60, '特別休暇'
        elif x == '免':
            a, b, c, d, e, f = 8, 30, 17, 15, 60, '職専免'
        elif x == '夏':
            a, b, c, d, e, f = 8, 30, 17, 15, 60, '夏季休暇'
        elif x == '代':
            a, b, c, d, e, f = None, None, None, None, None, '代休'
        elif x == '張':
            a, b, c, d, e, f = 8, 30, 17, 15, 60, '出張'
        else:
            a, b, c, d, e, f = None, None, None, None, None, None
        print(a,b,c,d,e,f)

        sheet_2.cell(row=start+i, column=4).value = a #8時
        sheet_2.cell(row=start+i, column=6).value = b #30分
        sheet_2.cell(row=start+i, column=8).value = c #17時
        sheet_2.cell(row=start+i, column=10).value = d #15分
        sheet_2.cell(row=start+i, column=13).value = e #60分（休憩時間）
        sheet_2.cell(row=start+i, column=32).value = f #コメント

    print('OK')
    st.write('Excelへの書き込みが完了しました。')
    

    wb.save(filepath_2)
    wb.close()
    st.download_button("Download", data=filepath_2, mime='xlsx', file_name=f'労働時間申請書_{year}_{month}.xlsx')

#     output = BytesIO()
#     workbook = xlsxwriter.Workbook(output, {'in_memory': True})
#     worksheet = workbook.add_worksheet()

#     worksheet.write('A1', 'Hello')
#     workbook.close()

#     st.download_button(
#         label="Download Excel workbook",
#         data=output.getvalue(),
#         file_name="workbook.xlsx",
#         mime="application/vnd.ms-excel"
# )
    # csv = df.to_csv(index=False, encoding='utf-8-sig')
    # data = open(filepath_2, 'rb').read()
    # b64 = base64.b64encode(data.encode('utf-8-sig')).decode()
    # href = f'<a href="data:application/octet-stream;base64,{b64}" download="result.xlsx">download</a>'
    # st.markdown(f"ダウンロードする {href}", unsafe_allow_html=True)
    # xlsx_exporter = px.Workbook()
    # sheet = xlsx_exporter.active
    # xlsx_exporter.save(filepath_2)
    # data = BytesIO(filepath_2.read())
    # b64 = base64.b64encode(data).decode('utf-8-sig')
    # href = f'<a href="data:application/octet-stream;base64,{b64}" download="result.xlsx">download</a>'
    # st.markdown(href, unsafe_allow_html=True)
    # xlsx_exporter.close()

    # st.download_button("Download",
    #     data=data,
    #     mime='xlsx',
    #     file_name="name_of_file.xlsx")
